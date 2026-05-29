# lib/export_excel.py
import sys
import json
import openpyxl
import os

def export_excel(input_json_path, template_path, output_path):
    try:
        # Load clients data
        with open(input_json_path, 'r') as f:
            clients = json.load(f)
            
        # Open the template workbook
        wb = openpyxl.load_workbook(template_path)
        
        # Ensure "Data Entry" sheet exists
        if "Data Entry" not in wb.sheetnames:
            raise Exception("Template missing 'Data Entry' sheet")
            
        sheet = wb["Data Entry"]
        
        # Clear existing data rows (from row 6 to max_row)
        # We don't delete rows entirely to preserve formatting, we just clear values
        for row_idx in range(6, max(6, sheet.max_row + 1)):
            for col_idx in range(1, 46):
                sheet.cell(row=row_idx, column=col_idx).value = None

        # Dimension to column offset mapping
        dimensions_map = {
            "Life Overall": 4,
            "Standard of Living": 7,
            "Health": 10,
            "Achieving in Life": 13,
            "Personal Relationships": 16,
            "Safety": 19,
            "Community": 22,
            "Future Security": 25,
            "Financial Worry": 28,
            "Self-Confidence": 31,
            "Voice & Agency": 34,
            "Work Readiness": 37,
            "Career Confidence": 40,
            "Skills Awareness": 43
        }
        
        # Populate the template
        start_row = 6
        for i, client in enumerate(clients):
            row_idx = start_row + i
            
            # Basic info
            sheet.cell(row=row_idx, column=1).value = client.get("intakeId")
            sheet.cell(row=row_idx, column=2).value = client.get("id")
            sheet.cell(row=row_idx, column=3).value = client.get("cohort")
            
            # Scores
            scores = client.get("scores", {})
            baseline = scores.get("Baseline", {}) or {}
            mo3 = scores.get("3mo", {}) or {}
            mo6 = scores.get("6mo", {}) or {}
            
            for dim, offset in dimensions_map.items():
                if dim in baseline and baseline[dim] is not None:
                    sheet.cell(row=row_idx, column=offset).value = baseline[dim]
                if dim in mo3 and mo3[dim] is not None:
                    sheet.cell(row=row_idx, column=offset + 1).value = mo3[dim]
                if mo6 and dim in mo6 and mo6[dim] is not None:
                    sheet.cell(row=row_idx, column=offset + 2).value = mo6[dim]
                    
        # Save to output path
        wb.save(output_path)
        
        return {"ok": True, "file": output_path}
        
    except Exception as e:
        return {"error": str(e)}

if __name__ == "__main__":
    if len(sys.argv) < 4:
        print(json.dumps({"error": "Missing arguments. Usage: python3 export_excel.py <input_json> <template_xlsx> <output_xlsx>"}))
        sys.exit(1)
        
    input_json = sys.argv[1]
    template = sys.argv[2]
    output = sys.argv[3]
    
    result = export_excel(input_json, template, output)
    print(json.dumps(result))
